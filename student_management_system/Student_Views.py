from django.shortcuts import render, redirect
from django.http import HttpResponse
from app.models import Student_Notification, Student, Student_Feedback, Student_leave, Subject, Attendance, Attendance_Report, StudentResult, Session_Year, Note, StudyMaterial
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import io
import google.generativeai as genai
from django.conf import settings

@login_required(login_url='/')
def Home(request):
    student = Student.objects.get(admin=request.user)
    
    # Get all results for the student
    results = StudentResult.objects.filter(student_id=student)
    
    # Calculate attendance percentage
    attendance_reports = Attendance_Report.objects.filter(student_id=student)
    total_attendance = attendance_reports.count()
    present_attendance = attendance_reports.filter(status=1).count()
    attendance_percent = (present_attendance / total_attendance * 100) if total_attendance > 0 else 0
    
    # Get subject-wise marks
    subject_marks = []
    for result in results:
        total_mark = 0
        if result.ia1_mark: total_mark += result.ia1_mark
        if result.ia2_mark: total_mark += result.ia2_mark
        if result.midsem_mark: total_mark += result.midsem_mark
        if result.end_sem_mark: total_mark += result.end_sem_mark
        
        subject_marks.append({
            'subject': result.subject_id.name,
            'total_mark': total_mark
        })
    
    # Get notifications count
    notification_count = Student_Notification.objects.filter(student_id=student, status=0).count()
    
    # Get study materials count
    materials_count = StudyMaterial.objects.filter(subject__course=student.course_id).count()
    
    # Get notes count
    notes_count = Note.objects.filter(user=request.user).count()
    
    # Get leave applications status
    leave_pending = Student_leave.objects.filter(student_id=student, status=0).count()
    leave_approved = Student_leave.objects.filter(student_id=student, status=1).count()
    leave_rejected = Student_leave.objects.filter(student_id=student, status=2).count()
    
    context = {
        'attendance_percent': round(attendance_percent, 1),
        'subject_marks': subject_marks,
        'notification_count': notification_count,
        'materials_count': materials_count,
        'notes_count': notes_count,
        'leave_pending': leave_pending,
        'leave_approved': leave_approved,
        'leave_rejected': leave_rejected,
        'student': student,
    }
    return render(request, 'Student/home.html', context)

@login_required(login_url='/')
def STUDENT_NOTIFICATION(request):
    try:
        # Get the Student object linked to the logged-in user
        student_instance = Student.objects.get(admin=request.user)
        # Filter notifications for that specific student, ordered by creation date
        notifications = Student_Notification.objects.filter(student_id=student_instance).order_by('-created_at')
        context = {'notification': notifications}
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        # Pass an empty list to the template to avoid errors if profile is missing
        context = {'notification': []}
        # Optionally, redirect to a home or error page:
        # return redirect('Home') # Assuming 'Home' is the name of your student dashboard URL
    
    return render(request, 'Student/notification.html', context)

@login_required(login_url='/')
def STUDENT_NOTIFICATION_MARK_AS_DONE(request, status):
    try:
        notification = Student_Notification.objects.get(id=status, student_id__admin=request.user) # Ensure student owns notification
        notification.status = 1
        notification.save()
        messages.success(request, 'Notification marked as read.')
    except Student_Notification.DoesNotExist:
        messages.error(request, 'Notification not found or access denied.')
    return redirect('student_notification') # Name of the URL for student to view notifications

@login_required(login_url='/')
def STUDENT_FEEDBACK(request):
    student_id = Student.objects.get(admin=request.user.id)
    feedback_history = Student_Feedback.objects.filter(student_id=student_id).order_by('-id')
    context = {"feedback_history": feedback_history}
    return render(request, 'Student/feedback.html', context)

@login_required(login_url='/')
def STUDENT_FEEDBACK_SAVE(request):
    if request.method == "POST":
        feedback = request.POST.get('feedback')
        student = Student.objects.get(admin=request.user.id)
        feedbacks = Student_Feedback(student_id=student, feedback=feedback, feedback_reply="")
        feedbacks.save()
        messages.success(request, 'Feedback Successfully Sent!')
        return redirect('student_feedback')
    return redirect('student_feedback')

@login_required(login_url='/')
def STUDENT_LEAVE(request):
    student = Student.objects.get(admin=request.user.id)
    student_leave_history = Student_leave.objects.filter(student_id=student)
    context = {'student_leave_history': student_leave_history}
    return render(request, 'Student/apply_leave.html', context)

@login_required(login_url='/')
def STUDENT_LEAVE_SAVE(request):
    if request.method == "POST":
        leave_date = request.POST.get('leave_date')
        leave_message = request.POST.get('leave_message')

        if not leave_date:
            messages.error(request, 'Please provide a valid leave date.')
            return redirect('student_leave')
        if not leave_message or leave_message.strip() == '':
            messages.error(request, 'Please provide a reason for the leave.')
            return redirect('student_leave')

        student_id = Student.objects.get(admin=request.user.id)
        student_leave = Student_leave(student_id=student_id, date=leave_date, message=leave_message)
        student_leave.save()
        messages.success(request, 'Leave Application Successfully Sent!')
        return redirect('student_leave')
    return redirect('student_leave')

@login_required(login_url='/')
def STUDENT_VIEW_ATTENDANCE(request):
    student = Student.objects.get(admin=request.user.id)
    subjects = Subject.objects.filter(course=student.course_id)
    session_year = Session_Year.objects.all()
    action = request.GET.get('action')
    get_subject = None
    get_session_year = None
    attendance_report = None

    if action == 'show_attendance' and request.method == "POST":
        subject_id = request.POST.get('subject_id')
        session_year_id = request.POST.get('session_year_id')
        try:
            if subject_id and subject_id.isdigit() and session_year_id and session_year_id.isdigit():
                get_subject = Subject.objects.get(id=int(subject_id))
                get_session_year = Session_Year.objects.get(id=int(session_year_id))
                if get_subject.course != student.course_id or get_session_year.id != student.session_year_id.id:
                    messages.error(request, "You are not authorized to view attendance for this subject or session year.")
                    return redirect('student_view_attendance')
                attendance_report = Attendance_Report.objects.filter(
                    student_id=student,
                    attendance_id__subject_id=get_subject,
                    attendance_id__session_year_id=get_session_year
                ).order_by('attendance_id__attendance_data')
            else:
                messages.error(request, "Please select a valid subject and session year.")
                return redirect('student_view_attendance')
        except Subject.DoesNotExist:
            messages.error(request, "Selected subject does not exist.")
            return redirect('student_view_attendance')
        except Session_Year.DoesNotExist:
            messages.error(request, "Selected session year does not exist.")
            return redirect('student_view_attendance')

    if 'download' in request.GET and request.GET['download'] == 'excel':
        subject_id = request.GET.get('subject_id')
        session_year_id = request.GET.get('session_year_id')
        print(f"Download Excel triggered - subject_id: {subject_id}, session_year_id: {session_year_id}")

        try:
            get_subject = Subject.objects.get(id=int(subject_id))
            get_session_year = Session_Year.objects.get(id=int(session_year_id))
            if get_subject.course != student.course_id or get_session_year.id != student.session_year_id.id:
                messages.error(request, "You are not authorized to download attendance for this subject or session year.")
                return redirect('student_view_attendance')
            attendance_report = Attendance_Report.objects.filter(
                student_id=student,
                attendance_id__subject_id=get_subject,
                attendance_id__session_year_id=get_session_year.id
            ).order_by('attendance_id__attendance_data')
        except (Subject.DoesNotExist, Session_Year.DoesNotExist, ValueError) as e:
            messages.error(request, "Invalid subject or session year selected.")
            return redirect('student_view_attendance')

        if attendance_report.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["Date", "Status"])
            for report in attendance_report:
                ws.append([report.attendance_id.attendance_data.strftime('%Y-%m-%d'), "Present" if report.status == 1 else "Absent"])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=attendance_{get_subject.name}_{get_session_year.session_start}_to_{get_session_year.session_end}.xlsx'
            wb.save(response)
            print("Excel file generated and sent for student")
            return response
        else:
            messages.error(request, "No attendance data available to export.")
            return redirect('student_view_attendance')

    if 'download_all' in request.GET and request.GET['download_all'] == 'excel':
        session_year_id = request.GET.get('session_year_id')
        print(f"Download All Subjects Excel triggered - session_year_id: {session_year_id}")

        try:
            get_session_year = Session_Year.objects.get(id=int(session_year_id))
            if get_session_year.id != student.session_year_id.id:
                messages.error(request, "You are not authorized to download attendance for this session year.")
                return redirect('student_view_attendance')
            all_subjects = Subject.objects.filter(course=student.course_id)
            attendance_reports = Attendance_Report.objects.filter(
                student_id=student,
                attendance_id__session_year_id=get_session_year.id,
                attendance_id__subject_id__in=all_subjects
            ).order_by('attendance_id__subject_id__name', 'attendance_id__attendance_data')
        except (Session_Year.DoesNotExist, ValueError) as e:
            messages.error(request, "Invalid session year selected.")
            return redirect('student_view_attendance')

        if attendance_reports.exists():
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove the default sheet

            for subject in all_subjects:
                subject_reports = attendance_reports.filter(attendance_id__subject_id=subject)
                if subject_reports.exists():
                    ws = wb.create_sheet(title=subject.name)
                    ws.append(["Date", "Status"])
                    for report in subject_reports:
                        ws.append([report.attendance_id.attendance_data.strftime('%Y-%m-%d'), "Present" if report.status == 1 else "Absent"])

            if not wb.sheetnames:
                messages.error(request, "No attendance data available to export for any subjects.")
                return redirect('student_view_attendance')

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=attendance_all_subjects_{get_session_year.session_start}_to_{get_session_year.session_end}.xlsx'
            wb.save(response)
            print("Excel file with all subjects generated and sent for student")
            return response
        else:
            messages.error(request, "No attendance data available to export for any subjects.")
            return redirect('student_view_attendance')

    context = {
        'subjects': subjects,
        'session_year': session_year,
        'action': action,
        'get_subject': get_subject,
        'get_session_year': get_session_year,
        'attendance_report': attendance_report,
    }
    return render(request, 'Student/view_attendance.html', context)

@login_required(login_url='/')
def VIEW_RESULT(request):
    student = Student.objects.get(admin=request.user.id)
    results = StudentResult.objects.filter(student_id=student)

    if 'download' in request.GET and request.GET['download'] == 'pdf':
        print("Download PDF triggered for student")
        # Create a PDF using reportlab
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        normal_style = styles['Normal']

        # Add title and student info
        elements.append(Paragraph("Student Results", title_style))
        elements.append(Paragraph(f"Student: {student.admin.first_name} {student.admin.last_name}", normal_style))
        elements.append(Paragraph(f"ID: {student.admin.id}", normal_style))
        elements.append(Paragraph("<br/><br/>", normal_style))  # Spacing

        # Table data
        data = [['Subject', 'Assignment', 'Exam', 'IA1', 'IA2', 'Attendance', 'Mid Sem', 'End Sem', 'CGPA']]
        for result in results:
            data.append([
                result.subject_id.name,
                str(result.assignment_mark) if result.assignment_mark is not None else '-',
                str(result.exam_mark) if result.exam_mark is not None else '-',
                str(result.ia1_mark) if result.ia1_mark is not None else '-',
                str(result.ia2_mark) if result.ia2_mark is not None else '-',
                str(result.attendance_mark) if result.attendance_mark is not None else '-',
                str(result.midsem_mark) if result.midsem_mark is not None else '-',
                str(result.end_sem_mark) if result.end_sem_mark is not None else '-',
                str(result.calculate_cgpa()) if result.calculate_cgpa() is not None else '-'
            ])

        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)

        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=my_results_{student.admin.username}.pdf'
        response.write(pdf)
        print("PDF file generated and sent for student")
        return response

    context = {'results': results}
    return render(request, 'Student/view_result.html', context)

@login_required(login_url='/')
def STUDENT_VIEW_MATERIALS(request):
    student = Student.objects.get(admin=request.user)
    # Get materials for subjects in student's course
    materials = StudyMaterial.objects.filter(subject__course=student.course_id).order_by('-uploaded_at')
    
    context = {
        'materials': materials,
    }
    return render(request, 'Student/view_materials.html', context)

# New views for student notes
@login_required(login_url='/')
def STUDENT_NOTES(request):
    student = Student.objects.get(admin=request.user)
    notes = Note.objects.filter(user=student.admin)
    context = {
        'notes': notes,
    }
    return render(request, 'Student/notes.html', context)

@login_required(login_url='/')
def STUDENT_CREATE_NOTE(request):
    if request.method == "POST":
        title = request.POST.get('title')
        content = request.POST.get('content')
        if title and content:
            note = Note(user=request.user, title=title, content=content)
            note.save()
            messages.success(request, 'Note created successfully!')
            return redirect('student_notes')
        else:
            messages.error(request, 'Title and content are required.')
    return render(request, 'Student/create_note.html')

@login_required(login_url='/')
def STUDENT_EDIT_NOTE(request, note_id):
    try:
        note = Note.objects.get(id=note_id, user=request.user)
    except Note.DoesNotExist:
        messages.error(request, 'Note not found.')
        return redirect('student_notes')
    if request.method == "POST":
        title = request.POST.get('title')
        content = request.POST.get('content')
        if title and content:
            note.title = title
            note.content = content
            note.save()
            messages.success(request, 'Note updated successfully!')
            return redirect('student_notes')
        else:
            messages.error(request, 'Title and content are required.')
    context = {
        'note': note,
    }
    return render(request, 'Student/edit_note.html', context)

@login_required(login_url='/')
def STUDENT_DELETE_NOTE(request, note_id):
    try:
        note = Note.objects.get(id=note_id, user=request.user)
        note.delete()
        messages.success(request, 'Note deleted successfully!')
    except Note.DoesNotExist:
        messages.error(request, 'Note not found.')
    return redirect('student_notes')


@login_required
def student_doubt_solver(request):
    context = {}
    
    if request.method == 'POST':
        question = request.POST.get('question')
        
        if not settings.GEMINI_API_KEY:
            messages.error(request, 'Gemini API key is not configured')
            return render(request, 'Student/doubt_solver.html', context)
            
        # Configure the Gemini API
        genai.configure(api_key=settings.GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-1.5-flash')  # Fixed model name
        
        # Enhanced prompt for student-friendly responses
        enhanced_prompt = f"""Please help explain this to a student in a very simple and easy-to-understand way:
1. Use simple language and avoid complex technical terms
2. Break down the explanation into small, digestible parts
3. Include relevant examples if possible
4. Use a friendly and encouraging tone
5. dont's try to bold any words or anything as ** looks ugly

Student's Question: {question}"""
        
        try:
            response = model.generate_content(enhanced_prompt)
            context['answer'] = response.text
        except Exception as e:
            messages.error(request, f'Error generating response: {str(e)}')
    
    return render(request, 'Student/doubt_solver.html', context)